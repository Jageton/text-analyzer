import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils.dataframe import dataframe_to_rows


class Output:
    @staticmethod
    def write_to_txt_file(file_path, file):
        """
        Записывает в txt файл
        :param file_path: путь файла для записи
        :param file: записываемый файл
        """

        try:
            with open(file_path, 'w', encoding='utf-8') as file2:
                file2.write(str(file))
        except TimeoutError:
            print('Истекло время ожидания')
        except Exception:
            print('Неизвестная ошибка!')

    @staticmethod
    def write_to_csv_file(file_path, csv_table):
        """
        Записывает в csv файл
        :param file_path: путь файла для записи
        :param csv_table: записываемая csv таблица
        """

        try:
            csv_table.to_csv(file_path)
        except TimeoutError:
            print('Истекло время ожидания')
        except Exception:
            print('Неизвестная ошибка!')

    @staticmethod
    def write_array_to_txt_file(file_path, array):
        """
        Записывает массив в файл
        :param file_path: путь файла для записи
        :param array: записываемый массив
        """

        try:
            with open(file_path, 'w', encoding='utf-8') as file2:
                for line in array:
                    file2.write(str(line) + "\n")
        except TimeoutError:
            print('Истекло время ожидания')
        except Exception:
            print('Неизвестная ошибка!')

    @staticmethod
    def write_to_xlsx_file(file_path, data_frame, clusters_dict, params):
        """
        Записывает данные в excel файл
        :param file_path: путь файла для записи
        :param data_frame: исходные данные
        :param clusters_dict: распределение по кластерам
        :param params: список параметров алгоритма
        """

        workbook = openpyxl.Workbook()
        ws_source_data = workbook.active
        ws_source_data.title = 'Исходные данные'

        rows = dataframe_to_rows(df=data_frame, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_source_data.cell(row=r_idx, column=c_idx, value=value)
                cell.alignment = Alignment(horizontal='center', vertical='center')

        Output._format_sheet(ws_source_data)

        ws_result = workbook.create_sheet('Результат')

        gl_row_idx = 1
        if 'alg_name' in params:
            cell = ws_result.cell(row=gl_row_idx, column=1, value='Алгоритм - {:s}'.format(params.pop('alg_name')))
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(bold=True)
            ws_result.merge_cells('A{:d}:B{:d}'.format(gl_row_idx, gl_row_idx))
            gl_row_idx += 2

        cell = ws_result.cell(row=gl_row_idx, column=1, value='Параметры алгоритма')
        ws_result.merge_cells('A{:d}:B{:d}'.format(gl_row_idx, gl_row_idx))
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(bold=True)
        for r_idx, (key, value) in enumerate(params.items(), 1):
            gl_row_idx += 1
            cell = ws_result.cell(row=gl_row_idx, column=1, value=key)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell = ws_result.cell(row=gl_row_idx, column=2, value=value)
            cell.alignment = Alignment(horizontal='center', vertical='center')

        gl_row_idx += 2
        for c_idx, (key, values) in enumerate(clusters_dict.items(), 1):
            cell = ws_result.cell(row=gl_row_idx, column=c_idx, value=key)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(bold=True)
            for r_idx, (k, v) in enumerate(values.items(), 1):
                cell = ws_result.cell(row=r_idx + gl_row_idx, column=c_idx,
                                      value='{:s}: {:s}'.format(str(k), Output._points_to_str(v)))
                cell.alignment = Alignment(horizontal='center', vertical='center')

        Output._format_sheet(ws_result)

        workbook.save(file_path)
        workbook.close()

    @staticmethod
    def _points_to_str(list):
        return '(%s)' % ';'.join(map(lambda x: str(x), list))

    @staticmethod
    def _format_sheet(worksheet):
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value or "")) for cell in column_cells) + 2
            cell = column_cells[0]
            letter = cell.column_letter if hasattr(cell, 'column_letter') else cell.coordinate[0:1]
            worksheet.column_dimensions[letter].width = length
